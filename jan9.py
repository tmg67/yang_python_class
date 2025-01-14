import openpyxl

#create a messy dataset in Excel for demonstration
def create_messy_dataset(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sales data"

    #add messy data
    sheet.append(["date", "product", "quantity", "price"])
    sheet.append(["2025-01-05", "laptop",3, 1200.50])
    sheet.append(["","mouse","",25.99])#missing date and quantity
    sheet.append(["2025-01-05", "laptop", 3, 1200.50])#duplicate row
    sheet.append(["2025-01-06", "keyboard", 2, None])#missing place

    workbook.save(file_path)
    print("messy dataset created")
    #clean the data set
def clean_dataset(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    cleaned_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #skips rows with missing data
        if None in row or "" in row:
            continue
        #remove duplicates
        if row not in cleaned_data:
            cleaned_data.append(row)

    #clear the sheet and write cleaned data
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    #re-add headers
    headers = ["date","product","quantity","price"]
    sheet.append(headers)
    #add cleaned data
    for row in cleaned_data:
        sheet.append(row)

    workbook.save(file_path)
    print("dataset cleaned")

#example usage
file_path = "messy_sales_data.xlsx"
create_messy_dataset(file_path)
clean_dataset(file_path)